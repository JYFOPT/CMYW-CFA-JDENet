import os
os.environ["PYTORCH_CUDA_ALLOC_CONF"] = "max_split_size_mb:128"
import torch.optim
from torch.utils.tensorboard import SummaryWriter
import os
from argparse import ArgumentParser
import numpy as np
from tqdm.autonotebook import tqdm
from src import  summary_utils, utils, mymodels
from place2.SIED_CMYW_GT_PatchDataset import CMYW_RGB_PatchDataset, RandomFlipBoth

from torch.utils.data import DataLoader
from torchvision.utils import save_image
import time
import torch.nn as nn
import sys
import xlwt
from openpyxl import load_workbook
from skimage.metrics import structural_similarity as ssim
from skimage.color import rgb2lab
def compute_delta_e(rgb1, rgb2):
    """
    rgb1, rgb2: numpy arrays of shape (H, W, 3) in range [0, 1]
    returns: mean Delta E (CIELAB Euclidean distance)
    """
    lab1 = rgb2lab(rgb1)
    lab2 = rgb2lab(rgb2)
    diff = np.linalg.norm(lab1 - lab2, axis=2)
    return diff.mean()
def delta_e_loss(pred_rgb, target_rgb):
    """
    pred_rgb, target_rgb: (B,3,H,W) in [0,1]
    returns: scalar loss (mean ΔE over all pixels)
    """
    # 将 tensor 转为 numpy 再转 lab 效率低，可用自定义 CUDA 或近似公式
    # 这里提供一个简化的可微 LAB 转换（基于 PyTorch）
    def rgb_to_lab_torch(img):
        # img: (B,3,H,W) in [0,1]
        # 转换到线性 RGB
        rgb_linear = torch.where(img > 0.04045, ((img + 0.055) / 1.055) ** 2.4, img / 12.92)
        # RGB to XYZ (D65)
        matrix = torch.tensor([[0.4124564, 0.3575761, 0.1804375],
                               [0.2126729, 0.7151522, 0.0721750],
                               [0.0193339, 0.1191920, 0.9503041]], device=img.device)
        xyz = torch.einsum('ij,bjhw->bihw', matrix, rgb_linear)
        # Normalize for D65 white point
        xyz_norm = xyz / torch.tensor([0.95047, 1.0, 1.08883], device=img.device).view(3,1,1)
        # XYZ to Lab
        epsilon = 0.008856
        kappa = 903.3
        f = torch.where(xyz_norm > epsilon, xyz_norm ** (1/3), (kappa * xyz_norm + 16) / 116)
        L = 116 * f[1] - 16
        a = 500 * (f[0] - f[1])
        b = 200 * (f[1] - f[2])
        lab = torch.stack([L, a, b], dim=1)
        return lab

    pred_lab = rgb_to_lab_torch(pred_rgb)
    target_lab = rgb_to_lab_torch(target_rgb)
    delta_e = torch.sqrt(torch.sum((pred_lab - target_lab) ** 2, dim=1) + 1e-6)  # (B,H,W)
    return delta_e.mean()

#记得修改网络输出的通道数，改为3

print('torch.cuda.device_count()',torch.cuda.device_count())
# exit()
os.environ['CUBLAS_WORKSPACE_CONFIG'] = ':4096:8'
utils.seed(num=123,deterministic=True)
if torch.cuda.device_count() ==5:
    device = torch.device('cuda:4' if torch.cuda.is_available() else 'cpu')
elif torch.cuda.device_count()==4:
    device = torch.device('cuda:3' if torch.cuda.is_available() else 'cpu')
elif torch.cuda.device_count()==3:
    device = torch.device('cuda:2' if torch.cuda.is_available() else 'cpu')
else:
    device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
    
print('device',device)
# exit(0)
def main(args):
    print(f"Actual batch_size: {args.batch_size}")
    start_time=time.time()
    save_root_txt='../{:s}/ckpt/{:s}/{:s}/set.txt'.format(args.save_dir,
                                                          args.shutter,
                                                          args.interp)
    fo = open(save_root_txt, 'a')
    fo.write('epoch:{}, '.format(args.max_epochs))
    fo.write('batch_size:{}, '.format(args.batch_size))
    fo.write('cfa_size:{}, '.format(args.cfa_size))
    fo.write('slr:{}, '.format(args.slr))
    fo.write('mlr:{}, '.format(args.mlr))
    fo.write('shutter:{}, \n'.format(args.shutter))
    fo.write('interpolation:{}, '.format(args.interp))
    fo.write('decoder:{}, '.format(args.decoder))
    fo.write('proxy function:{}, \n'.format(args.init))
    fo.write('bool_noise:{}, '.format(args.bool_noise))
    fo.write('noise_std:{}, '.format(args.noise_std))
    fo.write('loss:{}, '.format(args.loss))
    fo.write('vis_interval:{}, \n'.format(args.vis_interval))
    fo.write('w_zhi:{}, '.format(args.w_zhi))
    fo.write('alpha:{}, '.format(args.alpha))
    fo.write('gaussian_weight_size:{}, \n'.format(args.gaussian_weight_size))
    fo.write('datasset:{}, \n'.format(args.root))
    fo.close()
    # torch.cuda.empty_cache() #清除pytorch缓存

    args.slr = float(args.slr) #slr--shutter lr
    args.mlr = float(args.mlr) #mlr--model lr


    if not os.path.exists(args.save_dir):
        os.makedirs('{:s}/images'.format(args.save_dir))
        os.makedirs('{:s}/ckpt'.format(args.save_dir))

    if not os.path.exists(args.log_dir):
        os.makedirs(args.log_dir)
    
    shutter=mymodels.define_myshutter(args.shutter,args)


    decoder=mymodels.define_mydecoder(args.decoder,args)
    model=mymodels.define_mymodel(shutter,decoder,args,get_coded=False)
    # print('model',model)
    # model.cuda() #原版
    """
    #查看有哪些需要学习的参数
    # print(model.parameters())
    # for name ,parameters in model.shutter.named_parameters():
        # print(name,':',parameters.size())
    
    for p in model.shutter.named_parameters():
        print('parameters',p) #输出 有哪些可学习参数

    """
    mylogs_dir='{:s}/{:s}/{:s}/{:s}/'.format(args.log_dir,args.shutter,args.interp,args.init)
    print(args.log_dir)
    writer = SummaryWriter(log_dir=mylogs_dir)
    optim = utils.mydefine_optim(model, args)

    # # SIED数据

    # ---------- 训练集（目录模式，按坐标配对） ----------
    train_cmyw_dir = r'../data/sied_sony_0.01_cmyw_train_data/split/train/patches/noisy_patches'
    train_gt_dir = r'../data/sied_sony_0.01_cmyw_train_data/split/train/patches/gt_patches_jpg'
    dataset_train = CMYW_RGB_PatchDataset(
        cmyw_path=train_cmyw_dir,
        gt_path=train_gt_dir,
        transform=RandomFlipBoth(),
        crop_size=None  # 已是 128×128，无需裁剪
    )

    # ---------- 验证集（若也有类似问题，可同样处理） ----------
    val_cmyw_dir = r'../data/sied_sony_0.01_cmyw_train_data/split/val/noisy'
    val_gt_dir = r'../data/sied_sony_0.01_cmyw_train_data/split/val/gt'
    dataset_val = CMYW_RGB_PatchDataset(
        cmyw_path=val_cmyw_dir,
        gt_path=val_gt_dir,
        transform=None,
        crop_size=None
    )
    print(f"训练集样本数: {len(dataset_train)}")
    print(f"验证集样本数: {len(dataset_val)}")

    val_dataloader=DataLoader(dataset_val,
                      batch_size=1,
                      num_workers=args.num_workers,
                      shuffle=False,
                    )

    loss_fn = utils.define_loss(args)

    best_val_psnr = 0
    best_val_ssim = 0  # SSIM 越高越好
    best_val_delta_e = float('inf')  # ΔE 越低越好
    psnr_mean_list=[]
    # print('len(train_dataloader)',len(train_dataloader)) #48 注意dataloader的长度和batch_size有关系，长度等于len(datasets_train)/batch_size向上取整
    print('len(val_dataloader)',len(val_dataloader)) #48
    # if torch.cuda.device_count() > 1:  # 检查电脑是否有多块GPU
        # print(f"Let's use {torch.cuda.device_count()} GPUs!")
        # model = nn.DataParallel(model)  # 将模型对象转变为多GPU并行运算的模型
    model.to(device)
    snapshot = '../{:s}/ckpt/{:s}/{:s}/{:s}/{:s}_epoch.pth'.format(args.snapshot, args.shutter, args.interp,args.init, args.test_epoch)
    continue_epoch=0
    if args.resume==True:
        print("从检查点处开始训练")
        print('snapshot',snapshot)
        checkpoint = torch.load(snapshot)
        try:
            model.load_state_dict(checkpoint['model_state_dict'])
        except KeyError:
            model.load_state_dict(checkpoint)
        continue_epoch=args.continue_epoch
    # print('model.shutter.parameters',model.shutter.parameters)
    # print('model.decoder.parameters',model.decoder.parameters)
    total_steps = 1 #原版为0
    for j in tqdm(range(continue_epoch, args.max_epochs)):


        # # ----如果数据集数量太大， 每个 epoch 重新随机选择 1/10 训练样本 ----
        num_train = len(dataset_train)
        subset_size = max(1, num_train // 10)
        indices = torch.randperm(num_train)[:subset_size].tolist()
        sampler = torch.utils.data.SubsetRandomSampler(indices)
        train_dataloader = DataLoader(dataset_train,
                                      batch_size=args.batch_size,
                                      sampler=sampler,
                                      num_workers=args.num_workers,
                                      )
        print('len(train_dataloader)', len(train_dataloader))  # 48


        model.train()
        for step, (model_input, gt) in enumerate(tqdm(train_dataloader)):
            # print('step',step)
            # print('model_input.shape',model_input.shape) #orch.Size([1, 4, 186, 317])
            # print('gt.shape',gt.shape) #torch.Size([1, 4, 186, 317])
            model_input = model_input.to(device)
            gt = gt.to(device)
            gt_rgb=gt[:,:3,:,:]


            restored,cfa_full= model(model_input,train=True)#修改版本
            cfa_img=cfa_full[: 3]
            # 原有重建损失（L2）
            recon_loss = loss_fn(restored, gt_rgb)
            # 颜色损失
            color_loss = delta_e_loss(restored, gt_rgb)
            # 设置颜色损失权重（可调参数）
            lambda_color = 1e-4
            total_loss = recon_loss + lambda_color * color_loss
            print(f'recon_loss: {recon_loss.item():.4f}, color_loss: {color_loss.item():.4f}, total: {total_loss.item():.4f}')
            optim.zero_grad()
            total_loss.backward()
            # 梯度裁剪，防止梯度爆炸
            torch.nn.utils.clip_grad_norm_(model.parameters(), max_norm=1.0)  # 可调参数
            optim.step()
            """
            if ((total_steps ) % args.save_loss == 0):
                writer.add_scalar('train_loss', train_loss, total_steps )
                print('保存训练损失')
            total_steps += 1
            """
        # """
        if (j+ 1) % args.save_model_interval == 0 or (j + 1) == args.max_epochs:
            save_dir='../{:s}/ckpt/{:s}/{:s}/{:s}/{:d}_{:d}_epoch.pth'.format(args.save_dir, args.shutter,args.interp,args.init,args.code,j+1)
            utils.save_mychkpt(model,optim,save_dir) #本篇论文自带
            # utils.save_myckpt(save_dir, [('model', model)], [('optimizer', optim)], j+1)  # PConv
        # """
        if (j+1) % args.save_cfa_epoch==0:
            save_cfa_dir='../{:s}/{:s}/{:s}/{:s}/{:d}_{:d}_epoch.png'.format(args.save_cfa_root,
                                                                             args.shutter,
                                                                             args.interp,
                                                                             args.init,
                                                                             args.code,
                                                                             j+1)
            for i in range(args.cfa_size[0]):
                for k in range(args.cfa_size[1]):
                    if cfa_img[0, i, k] == 1 and cfa_img[1, i, k] == 0 and cfa_img[2, i, k] == 0:
                        cfa_img[0, i, k] = 0
                        cfa_img[1, i, k] = 1
                        cfa_img[2, i, k] = 1
                    if cfa_img[0, i, k] == 0 and cfa_img[1, i, k] == 1 and cfa_img[2, i, k] == 0:
                        cfa_img[0, i, k] = 1
                        cfa_img[1, i, k] = 0
                        cfa_img[2, i, k] = 1
                    if cfa_img[0, i, k] == 0 and cfa_img[1, i, k] == 0 and cfa_img[2, i, k] == 1:
                        cfa_img[0, i, k] = 1
                        cfa_img[1, i, k] = 1
                        cfa_img[2, i, k] = 0
                    if cfa_img[0, i, k] == 0 and cfa_img[1, i, k] == 0 and cfa_img[2, i, k] == 0:
                        cfa_img[0, i, k] = 1
                        cfa_img[1, i, k] = 1
                        cfa_img[2, i, k] = 1
            save_image(cfa_img[:,:args.cfa_size[0],
                       :args.cfa_size[1]],
                       save_cfa_dir)
            # exit(0)
        psnr_mean_list = []
        ssim_mean_list = []
        delta_e_mean_list = []
        if (j + 1) % args.vis_interval == 0:
            if val_dataloader is not None:
                with torch.no_grad():
                    model.eval()
                    torch.cuda.empty_cache()
                    val_psnrs = []
                    val_ssims = []
                    val_delta_es = []

                    for (model_input, gt) in tqdm(val_dataloader):
                        model_input = model_input.to(device)
                        gt = gt.to(device)
                        gt_rgb = gt[:, :3, :, :]  # (1, 3, H, W)
                        restored, cfa_img = model(model_input, train=False)
                        # 检查模型输出范围
                        print(f"输出范围: [{restored.min():.3f}, {restored.max():.3f}]")
                        print(f"GT范围: [{gt_rgb.min():.3f}, {gt_rgb.max():.3f}]")

                        # 转为 numpy (H, W, 3) 用于 SSIM 和 ΔE
                        restored_np = restored.squeeze(0).permute(1, 2, 0).cpu().numpy()
                        gt_np = gt_rgb.squeeze(0).permute(1, 2, 0).cpu().numpy()
                        restored_np = np.clip(restored_np, 0, 1)
                        gt_np = np.clip(gt_np, 0, 1)

                        # PSNR（沿用原函数）
                        psnr = summary_utils.get_psnr(restored, gt_rgb)

                        # SSIM
                        ssim_val = ssim(restored_np, gt_np, channel_axis=2, data_range=1.0)

                        # ΔE
                        delta_e = compute_delta_e(restored_np, gt_np)

                        val_psnrs.append(psnr)
                        val_ssims.append(ssim_val)
                        val_delta_es.append(delta_e)

                    # 计算 epoch 平均值
                    mean_psnr = np.mean(val_psnrs)
                    mean_ssim = np.mean(val_ssims)
                    mean_delta_e = np.mean(val_delta_es)

                    # 打印到终端
                    print(f'Epoch {j + 1}: PSNR = {mean_psnr:.4f}, SSIM = {mean_ssim:.4f}, ΔE = {mean_delta_e:.4f}')

                    # 写入临时日志文件（每个 epoch 都写）
                    with open(save_root_txt, 'a') as fo:
                        fo.write(f'Epoch {j + 1}:\n')
                        fo.write(f'  PSNR: {mean_psnr:.4f}\n')
                        fo.write(f'  SSIM: {mean_ssim:.4f}\n')
                        fo.write(f'  ΔE  : {mean_delta_e:.4f}\n')
                        fo.write('---\n')

                    # 更新最佳值
                    if mean_psnr > best_val_psnr:
                        best_val_psnr = mean_psnr
                    if mean_ssim > best_val_ssim:
                        best_val_ssim = mean_ssim
                    if mean_delta_e < best_val_delta_e:
                        best_val_delta_e = mean_delta_e

                    # 记录到全局列表（若需要输出整个训练曲线）
                    psnr_mean_list.append(mean_psnr)
                    ssim_mean_list.append(mean_ssim)  # 需提前定义
                    delta_e_mean_list.append(mean_delta_e)  # 需提前定义
    print('psnr_mean_list',psnr_mean_list,'--len(psnr_mean_list)',len(psnr_mean_list))
    print('best_val_psnr',best_val_psnr)

    writer.close()
    print("epochs:  ",args.max_epochs)
    print("batch_size:",args.batch_size)
    print("cfa_size:",args.cfa_size)
    print("slr:",args.slr)
    print("mlr:",args.mlr)
    print("shutter: ",args.shutter ) 
    print("interpolation: ",args.interp)
    print("decoder:",args.decoder)
    print('代理函数：',args.init)
    print('固定cfaw初始化值为：',args.w_zhi)
    print("alpha: ",args.alpha)
    print("gaussian_weight_size: ",args.gaussian_weight_size)
    print("是否加噪声:",args.bool_noise)
    print("噪声等级std:",args.noise_std)
    end_time=time.time()
    print(f"程序运行时间：{end_time-start_time:.4f}s")
    fo = open(save_root_txt, 'a')
    fo.write('程序运行时间:{:.4f}s \n'.format(end_time - start_time))
    fo.write('psnr_mean_list:\n{} {}\n'.format(psnr_mean_list, len(psnr_mean_list)))
    fo.write('best_val_psnr:{}\n\n'.format(best_val_psnr))
    fo.write('===== Final Best Metrics =====\n')
    fo.write(f'Best PSNR : {best_val_psnr:.4f}\n')
    fo.write(f'Best SSIM : {best_val_ssim:.4f}\n')
    fo.write(f'Best ΔE   : {best_val_delta_e:.4f}\n')
    fo.write('================================\n')
    fo.close()  # 确保最后关闭
    return best_val_psnr



    
    

if __name__ == '__main__':
    parser = ArgumentParser()
    parser.add_argument('--root', type=str, default='.')
    # parser.add_argument('--root', type=str, default='../canon_data/568_great_test') #568_great_test  568_great
    # parser.add_argument('--root', type=str, default='../canon_data/data_tiaocan')
    parser.add_argument('--snapshot', type=str,default='snapshots/default')
    parser.add_argument('--test_epoch', type=str, default='1_85')
    parser.add_argument('--continue_epoch', type=int, default=85) #如果从模型中间开始训练必须修改从第几个epoch开始训练
    parser.add_argument('--resume',
                    type=bool,
                    default=False,
                    help='是否需要从检查点处开始训练，如需要请修改，test_epoch,并且其他参数也务必相同')
    parser.add_argument('-b', '--block_size',
                    help='delimited list input for block size in format(格式中块大小的分隔列表输入) %,%,%',
                    default=[4,128,128]) #default=[4,324,487] default='8,512,512' 8,1359,2041
    # parser.add_argument('--reg', type=float, default=100.0, help='regularization on lpips loss(lpips损失的正规化)')
    parser.add_argument('--cfa_size', help='cfa的尺寸大小,只针对lcmyw有效%,%,%',
                    default=[4,4])
    parser.add_argument('--k', type=int, default=3) #原版default=3
    parser.add_argument('--p', type=float, default=1.0)
    parser.add_argument('--bool_noise',type=bool,default=False)
    parser.add_argument('--noise_std',type=float,default=0.01)
    parser.add_argument('--max_epochs', type=int, default=4500) #原版6000
    parser.add_argument('--mlr', help='model_lr', type=str, default='2e-4') #原版 default='5e-4'
    parser.add_argument('--slr', help='shutter_lr', type=str, default='1e-3') #原版 default='2e-4'
    parser.add_argument('--batch_size', type=int,default=1) #原版default=2
    parser.add_argument('--w_zhi', type=float,default=0.001)
    parser.add_argument('--alpha', type=float,default=2.5)
    parser.add_argument('--gaussian_weight_size', type=int,default=5)
    parser.add_argument('--num_workers', type=int, default=0) #原版1
    parser.add_argument('--interp', type=str, choices=['none',
                                                       'gaussian'],
                        default='none') #原版required=True, default='pconv'/
    parser.add_argument('--init', type=str, choices=['softmax_tau',
                                                     'gumbel_softmax_tau_1',
                                                     'gumbel_softmax_tau_2'],
                        default='gumbel_softmax_tau_2',
                help='针对可学习CMYW CFA的形式,反向传播代理函数问题,'
                        'softmax_tau:带温度系数的softmax函数,gama=2.5e-5,可去函数里面调整,tau根据迭代次数进行变化,'
                        'gumbel_softmax_tau_1:Gumbel_Softmax函数,tau为初始值1,原始GB-ST'
                        'gumbel_softmax_tau_2:Gumbel_Softmax,tau可变,根据迭代次数变化,噪声项含有独立温度系数')
    parser.add_argument('--loss', type=str, choices=['mpr', 'l1', 'l2_lpips', 'l2'], default='l2')
    parser.add_argument('--decoder', type=str,
                        choices=['unet', 'RDNet', 'dncnn', 'myunet','Uformer',
                                 'mysgnet','drunet','rlfn','nafnet',
                                 'mysgnet1','rrdbnet','rdunet','simdunet',
                                 'RestNet','RASTNet','DGAsimdunet','DGAGLMIFunet',
                                 'Res3unet','JDENet','DNF','DUJRENet'],
                        default='JDENet') #原版default='unet'
    parser.add_argument('--shutter', type=str, default='lcmyw')
    parser.add_argument('--save_dir', type=str,
                        default='snapshots/default')
    parser.add_argument('--log_dir', type=str, default='mylogs')
    parser.add_argument('--save_model_interval', type=int, default=1) #保存模型时间间隔default=50000
    parser.add_argument('--vis_interval', type=int, default=1)#垂直间隔,每几个epoch去验证集
    parser.add_argument('--save_loss', type=int, default=10000000000000000)#经过几个迭代保存一次loss
    parser.add_argument('--save_cfa_epoch', type=int, default=1)#经过几个epoch保存一次cfa图像
    parser.add_argument('--save_cfa_root', type=str, default='cfa_image')
    parser.add_argument('--code', type=int, default=1)
    args = parser.parse_args()
    # main(args)


    # """
    read_root_txt='../{:s}/ckpt/{:s}/{:s}/set_canshu.txt'.format(args.save_dir,
                                                                 args.shutter,
                                                                 args.interp)
    save_root_xlsx='../{:s}/ckpt/{:s}/{:s}/demo.xlsx'.format(args.save_dir,
                                                             args.shutter,
                                                             args.interp)
    #加载excel，注意路径要与脚本一致
    wb = load_workbook(save_root_xlsx)
    #激活excel表
    sheet = wb.active
    fo_main=open(read_root_txt,'r')
    read_lines=fo_main.readlines()
    print('read_lines',read_lines)
    print('len(read_lines)',len(read_lines))
    code=0
    for m  in tqdm(range(1,len(read_lines))):
    # print('m',m)
    # exit()
        cfa_size=[]
        parameter=read_lines[m].strip('\n').split(',')
        # print(parameter)
        # exit()
        args.code=int(parameter[0])
        args.batch_size=int(parameter[1])
        args.slr=parameter[2]
        args.alpha=float(parameter[3])
        args.w_zhi=float(parameter[4])
        args.gaussian_weight_size=int(parameter[5])
        args.mlr=parameter[6]
        cfa_sz=parameter[7].split('x')
        cfa_size.append(int(cfa_sz[0]))
        cfa_size.append(int(cfa_sz[1]))
        args.cfa_size=cfa_size

        best_psnr=main(args)

        sheet.cell(row=m+1, column=1).value = args.code
        sheet.cell(row=m+1, column=2).value = args.interp
        sheet.cell(row=m+1, column=3).value = args.max_epochs
        sheet.cell(row=m+1, column=4).value = args.batch_size
        sheet.cell(row=m+1, column=5).value = str(args.cfa_size)
        sheet.cell(row=m+1, column=6).value = 'AdamW'
        sheet.cell(row=m+1, column=7).value = float(args.slr)
        sheet.cell(row=m+1, column=8).value = float(args.mlr)
        sheet.cell(row=m+1, column=9).value = args.alpha
        sheet.cell(row=m+1, column=10).value = args.w_zhi
        sheet.cell(row=m+1, column=11).value = args.gaussian_weight_size
        sheet.cell(row=m+1, column=12).value = best_psnr
        sheet.cell(row=m+1, column=13).value = '无'
        wb.save(save_root_xlsx)
        print("数据写入成功!")
        print('11111')
    fo_main.close()
    # exit()
    # """
